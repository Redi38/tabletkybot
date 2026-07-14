import csv
import io
from datetime import datetime, timezone
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn
from openpyxl.worksheet.worksheet import Worksheet

from locales.texts import get_text


# ── Helper functions ─────────────────────────────────────────────────────
def _prepare_and_sort_records(records: list[tuple], user_tz_str: str) -> list[tuple]:
    """Helper function for parsing dates, converting to local time, and sorting records."""
    parsed_records = []

    try:
        user_tz = ZoneInfo(user_tz_str)
    except ZoneInfoNotFoundError:
        user_tz = ZoneInfo("Europe/Kyiv")

    for r in records:
        name, dosage, remaining_days, taken_at, status = r

        taken_dt = taken_at if isinstance(taken_at, datetime) else datetime.fromisoformat(str(taken_at))

        if taken_dt.tzinfo is None:
            taken_dt = taken_dt.replace(tzinfo=timezone.utc)

        local_dt = taken_dt.astimezone(user_tz)

        parsed_records.append((name, dosage, remaining_days, local_dt, status))

    # Sort by date descending (newest on top)
    return sorted(parsed_records, key=lambda x: x[3], reverse=True)


def _get_clean_status_text(status: str, lang: str) -> str:
    """Gets the status text and strips emojis from it."""
    raw_status_text = get_text(lang, "excel_status_taken") if status == "taken" else get_text(lang,
                                                                                              "excel_status_skipped")

    return raw_status_text.replace("✅", "").replace("❌", "").replace("⏭️", "").strip()


def _build_medicine_stats_sheet(wb, sorted_records: list[tuple], lang: str,
                                 header_font: Font, center: Alignment,
                                 border: Border) -> None:
    """Summary table: one row per medicine + intake % + autofilter."""
    ws = wb.create_sheet(title=get_text(lang, "excel_med_stats_sheet"))

    # ── Row 1: sheet title ─────────────────────────────────────────
    ws.merge_cells("A1:G1")
    c: Any = ws.cell(row=1, column=1)
    c.value = get_text(lang, "excel_med_stats_title")
    c.font = Font(name="Calibri", bold=True, size=13)
    c.alignment = center
    ws.row_dimensions[1].height = 24

    # ── Row 2: column headers ───────────────────────────────────────
    med_headers = [
        get_text(lang, "excel_h_med_name"),
        get_text(lang, "excel_h_med_dose"),
        get_text(lang, "excel_h_med_taken"),
        get_text(lang, "excel_h_med_missed"),
        get_text(lang, "excel_h_med_total"),
        get_text(lang, "excel_h_med_pct"),
        get_text(lang, "excel_h_med_last"),
    ]
    med_widths = [22, 14, 12, 12, 10, 14, 18]
    med_fill = PatternFill("solid", fgColor="4472C4")

    for col_idx, (hdr, width) in enumerate(zip(med_headers, med_widths), start=1):
        c = ws.cell(row=2, column=col_idx)
        c.value = hdr
        c.font = header_font
        c.fill = med_fill
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 20

    # ── Aggregation by medicine ───────────────────────────────────────
    med_stats: dict[tuple, dict] = {}  # (name, dosage) → {taken, missed, last_dt}

    for name, dosage, _, taken_dt, status in sorted_records:
        key = (name, dosage)
        if key not in med_stats:
            med_stats[key] = {"taken": 0, "missed": 0, "last_dt": None}
        if status == "taken":
            med_stats[key]["taken"] += 1
            if med_stats[key]["last_dt"] is None or taken_dt > med_stats[key]["last_dt"]:
                med_stats[key]["last_dt"] = taken_dt
        elif status in ["missed", "skipped"]:
            med_stats[key]["missed"] += 1

    # ── Data: one row per medicine ──────────────────────────────────────
    pct_good = PatternFill("solid", fgColor="C6EFCE")
    pct_mid  = PatternFill("solid", fgColor="FFEB9C")
    pct_bad  = PatternFill("solid", fgColor="FFC7CE")

    for row_idx, ((name, dosage), data) in enumerate(
        sorted(med_stats.items(), key=lambda x: x[0][0].lower()), start=3
    ):
        taken  = data["taken"]
        missed = data["missed"]
        total  = taken + missed
        pct    = round(taken / total * 100, 1) if total > 0 else 0.0
        last_str = data["last_dt"].strftime("%d.%m.%Y %H:%M") if data["last_dt"] else "—"

        for col_idx, value in enumerate([name, dosage, taken, missed, total, pct, last_str], start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = value
            c.alignment = center
            c.border = border
            if col_idx == 6:  # intake % — color indicator
                if pct >= 80:
                    c.fill = pct_good
                    c.font = Font(name="Calibri", bold=True, color="375623")
                elif pct >= 50:
                    c.fill = pct_mid
                    c.font = Font(name="Calibri", bold=True, color="7E6000")
                else:
                    c.fill = pct_bad
                    c.font = Font(name="Calibri", bold=True, color="9C0006")

    # ── Freeze pane ──────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

# ── Main report generators ──────────────────────────────────────────────
def create_excel_report(records: list[tuple], lang: str = "ua", user_name: str = "",
                        user_tz: str = "Europe/Kyiv") -> io.BytesIO:
    """Generates an Excel report of medicine intake (History + Statistics)."""
    wb = openpyxl.Workbook()
    ws = wb.active

    if not isinstance(ws, Worksheet):
        return io.BytesIO()

    ws.title = get_text(lang, "excel_title")

    # ── Styles ──────────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_bold = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_fill_taken = PatternFill("solid", fgColor="E2EFDA")  # green
    status_fill_skipped = PatternFill("solid", fgColor="FCE4D6")  # red

    stat_header_fill = PatternFill("solid", fgColor="B0E0E6")  # light blue for statistics

    sorted_records = _prepare_and_sort_records(records, user_tz)

    # ── Sheet title (Sheet 1) ──────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title_cell: Any = ws.cell(row=1, column=1)
    title_cell.value = f"{get_text(lang, 'excel_title')} – {datetime.now().strftime('%d.%m.%Y')}"
    title_cell.font = Font(name="Calibri", bold=True, size=13)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 24

    # ── Patient name ──────────────────────────────────────────────────────
    ws.merge_cells("A2:F2")
    patient_cell: Any = ws.cell(row=2, column=1)
    patient_cell.value = f"{get_text(lang, 'excel_patient')} {user_name}"
    patient_cell.font = Font(name="Calibri", bold=True, size=11)
    patient_cell.alignment = left_bold
    ws.row_dimensions[2].height = 20

    # ── Column headers ─────────────────────────────────────────────────────
    headers = [
        get_text(lang, "excel_h_num"),
        get_text(lang, "excel_h_name"),
        get_text(lang, "excel_h_dose"),
        get_text(lang, "excel_h_date"),
        get_text(lang, "excel_h_time"),
        get_text(lang, "excel_h_status")
    ]
    widths = [5, 22, 15, 14, 10, 14]

    cell: Any = None
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 20

    # ── Data (Sheet 1) ──────────────────────────────────────────────────────
    for row_idx, record in enumerate(sorted_records, start=4):
        name, dosage, remaining_days, taken_dt, status = record

        status_text = _get_clean_status_text(status, lang)
        fill = status_fill_taken if status == "taken" else status_fill_skipped

        row_data = [
            row_idx - 3,
            name,
            dosage,
            taken_dt.strftime("%d.%m.%Y"),
            taken_dt.strftime("%H:%M"),
            status_text,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = center
            cell.border = border
            if col_idx == 6:
                cell.fill = fill

    # ── AutoFilter on the log ──────────────────────────────────────────────
    last_data_row = 3 + len(sorted_records)
    ws.auto_filter.ref = f"A3:F{last_data_row}"

    for col_idx in range(6):  # Remove autofilter buttons from A to F (0 to 5)
        if col_idx != 1:    # Column B index
            col_filter = FilterColumn(colId=col_idx, hiddenButton=True, blank=False)
            ws.auto_filter.filterColumn.append(col_filter)

    ws.freeze_panes = "A4"

    # ── Statistics (Sheet 2) ────────────────────────────────────────────────
    ws_stats = wb.create_sheet(title=get_text(lang, "excel_stats_sheet"))

    stats_data = {
        "0-30": {"taken": 0, "missed": 0},
        "31-60": {"taken": 0, "missed": 0},
        "61-90": {"taken": 0, "missed": 0},
        "91-120": {"taken": 0, "missed": 0},
        "121-150": {"taken": 0, "missed": 0},
        "151-180": {"taken": 0, "missed": 0},
        "180+": {"taken": 0, "missed": 0},
    }

    now = datetime.now(sorted_records[0][3].tzinfo) if sorted_records else datetime.now()

    for r in sorted_records:
        taken_dt = r[3]
        status = r[4]
        days_ago = (now - taken_dt).days

        if status == "taken":
            s_key = "taken"
        elif status in ["missed", "skipped"]:
            s_key = "missed"
        else:
            continue

        if 0 <= days_ago <= 30:
            stats_data["0-30"][s_key] += 1
        elif 31 <= days_ago <= 60:
            stats_data["31-60"][s_key] += 1
        elif 61 <= days_ago <= 90:
            stats_data["61-90"][s_key] += 1
        elif 91 <= days_ago <= 120:
            stats_data["91-120"][s_key] += 1
        elif 121 <= days_ago <= 150:
            stats_data["121-150"][s_key] += 1
        elif 151 <= days_ago <= 180:
            stats_data["151-180"][s_key] += 1
        else:
            stats_data["180+"][s_key] += 1

    stat_headers = [
        get_text(lang, "excel_stats_period"), get_text(lang, "excel_days_0_30"),
        get_text(lang, "excel_days_31_60"), get_text(lang, "excel_days_61_90"),
        get_text(lang, "excel_days_91_120"), get_text(lang, "excel_days_121_150"),
        get_text(lang, "excel_days_151_180"), get_text(lang, "excel_days_180_plus")
    ]

    for col_idx, header in enumerate(stat_headers, start=1):
        cell = ws_stats.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.fill = stat_header_fill
        cell.alignment = center
        cell.border = border
        ws_stats.column_dimensions[get_column_letter(col_idx)].width = 15

    ws_stats.column_dimensions['A'].width = 18

    row_taken: list[Any] = [get_text(lang, "excel_pure_taken")]
    row_missed: list[Any] = [get_text(lang, "excel_pure_skipped")]

    for p_data in stats_data.values():
        row_taken.append(p_data["taken"])
        row_missed.append(p_data["missed"])

    for row_idx, row_data in enumerate([row_taken, row_missed], start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_stats.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = center
            cell.border = border

            if row_idx == 3 and col_idx > 1 and isinstance(value, int) and value > 0:
                cell.font = Font(color="FF0000", bold=True)

    # ── By medicine (Sheet 3) ────────────────────────────────────────────
    _build_medicine_stats_sheet(wb, sorted_records, lang, header_font, center, border)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_csv_report(records: list[tuple], lang: str = "ua", user_name: str = "",
                      user_tz: str = "Europe/Kyiv") -> io.BytesIO:
    """Generates a lightweight CSV report of medicine intake."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Patient information
    writer.writerow([f"{get_text(lang, 'excel_patient')} {user_name}"])
    writer.writerow([])

    # Column headers
    headers = [
        get_text(lang, "excel_h_num"), get_text(lang, "excel_h_name"),
        get_text(lang, "excel_h_dose"), get_text(lang, "excel_h_date"),
        get_text(lang, "excel_h_time"), get_text(lang, "excel_h_status")
    ]
    writer.writerow(headers)

    # Use the helper function to prepare the data, accounting for timezone
    sorted_records = _prepare_and_sort_records(records, user_tz)

    # Data
    for row_idx, record in enumerate(sorted_records, start=1):
        name, dosage, remaining_days, taken_dt, status = record

        status_text = _get_clean_status_text(status, lang)

        writer.writerow([
            row_idx,
            name,
            dosage,
            taken_dt.strftime("%d.%m.%Y"),
            taken_dt.strftime("%H:%M"),
            status_text
        ])

    buffer = io.BytesIO(output.getvalue().encode('utf-8'))
    buffer.seek(0)
    return buffer
