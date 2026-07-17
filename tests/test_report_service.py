"""
Tests for services/report_service.py.

Covers the pure data-prep/aggregation helpers directly, plus smoke tests
for the Excel/CSV report generators (making sure they don't crash and
produce non-empty, readable output).
"""
from datetime import datetime, timezone

import openpyxl
import pytest

from services import report_service


def _dt(y, m, d, h=0, mi=0, tz=None):
    return datetime(y, m, d, h, mi, tzinfo=tz)


# prepare_and_sort_records

def test_prepare_and_sort_records_sorts_newest_first():
    records = [
        ("Aspirin", "500mg", 5, _dt(2026, 1, 1, 8, 0, tz=timezone.utc), "taken"),
        ("Aspirin", "500mg", 5, _dt(2026, 1, 3, 8, 0, tz=timezone.utc), "taken"),
        ("Aspirin", "500mg", 5, _dt(2026, 1, 2, 8, 0, tz=timezone.utc), "taken"),
    ]

    result = report_service._prepare_and_sort_records(records, "Europe/Kyiv")

    dates = [r[3].date() for r in result]
    assert dates == sorted(dates, reverse=True)


def test_prepare_and_sort_records_converts_naive_datetime_as_utc():
    """Naive datetimes (no tzinfo) coming from the DB should be treated as UTC."""
    naive = _dt(2026, 1, 1, 12, 0)  # no tzinfo
    records = [("Med", "10mg", 0, naive, "taken")]

    result = report_service._prepare_and_sort_records(records, "Europe/Kyiv")

    # Europe/Kyiv is UTC+2 in January
    assert result[0][3].hour == 14
    assert result[0][3].tzinfo is not None


def test_prepare_and_sort_records_converts_aware_datetime_to_local_tz():
    aware = _dt(2026, 6, 1, 12, 0, tz=timezone.utc)
    records = [("Med", "10mg", 0, aware, "taken")]

    # Europe/Kyiv is UTC+3 in summer (DST)
    result = report_service._prepare_and_sort_records(records, "Europe/Kyiv")

    assert result[0][3].hour == 15


def test_prepare_and_sort_records_handles_iso_string_taken_at():
    """taken_at can arrive as an ISO string instead of a datetime object."""
    records = [("Med", "10mg", 0, "2026-01-01T12:00:00+00:00", "taken")]

    result = report_service._prepare_and_sort_records(records, "Europe/Kyiv")

    assert result[0][3].hour == 14


def test_prepare_and_sort_records_falls_back_to_kyiv_on_invalid_timezone():
    """An invalid/unknown IANA timezone string should not crash - falls back to Europe/Kyiv."""
    records = [("Med", "10mg", 0, _dt(2026, 1, 1, 12, 0, tz=timezone.utc), "taken")]

    result = report_service._prepare_and_sort_records(records, "Not/A_Real_Zone")

    assert result[0][3].hour == 14


def test_prepare_and_sort_records_preserves_all_fields():
    records = [("Ibuprofen", "200mg", 7, _dt(2026, 1, 1, 8, 0, tz=timezone.utc), "skipped")]

    result = report_service._prepare_and_sort_records(records, "Europe/Kyiv")

    name, dosage, remaining_days, taken_dt, status = result[0]
    assert name == "Ibuprofen"
    assert dosage == "200mg"
    assert remaining_days == 7
    assert status == "skipped"


# aggregate_medicine_stats / get_medicine_stats_summary

def test_aggregate_medicine_stats_counts_taken_and_missed():
    sorted_records = [
        ("Aspirin", "500mg", 5, _dt(2026, 1, 3, tz=timezone.utc), "taken"),
        ("Aspirin", "500mg", 5, _dt(2026, 1, 2, tz=timezone.utc), "skipped"),
        ("Aspirin", "500mg", 5, _dt(2026, 1, 1, tz=timezone.utc), "taken"),
    ]

    result = report_service._aggregate_medicine_stats(sorted_records)

    assert len(result) == 1
    stat = result[0]
    assert stat["taken"] == 2
    assert stat["missed"] == 1
    assert stat["total"] == 3
    assert stat["pct"] == pytest.approx(66.7, abs=0.1)


def test_aggregate_medicine_stats_treats_missed_and_skipped_the_same():
    sorted_records = [
        ("Med", "10mg", 0, _dt(2026, 1, 1, tz=timezone.utc), "missed"),
        ("Med", "10mg", 0, _dt(2026, 1, 2, tz=timezone.utc), "skipped"),
    ]

    result = report_service._aggregate_medicine_stats(sorted_records)

    assert result[0]["missed"] == 2
    assert result[0]["taken"] == 0


def test_aggregate_medicine_stats_last_dt_is_most_recent_taken_dose():
    sorted_records = [
        ("Med", "10mg", 0, _dt(2026, 1, 3, tz=timezone.utc), "taken"),
        ("Med", "10mg", 0, _dt(2026, 1, 5, tz=timezone.utc), "taken"),
        ("Med", "10mg", 0, _dt(2026, 1, 1, tz=timezone.utc), "skipped"),
    ]

    result = report_service._aggregate_medicine_stats(sorted_records)

    assert result[0]["last_dt"] == _dt(2026, 1, 5, tz=timezone.utc)


def test_aggregate_medicine_stats_no_taken_doses_last_dt_is_none():
    sorted_records = [("Med", "10mg", 0, _dt(2026, 1, 1, tz=timezone.utc), "skipped")]

    result = report_service._aggregate_medicine_stats(sorted_records)

    assert result[0]["last_dt"] is None
    assert result[0]["pct"] == 0.0


def test_aggregate_medicine_stats_separates_different_dosages_of_same_medicine():
    """(name, dosage) is the grouping key, so different dosages of the same med are separate rows."""
    sorted_records = [
        ("Aspirin", "500mg", 0, _dt(2026, 1, 1, tz=timezone.utc), "taken"),
        ("Aspirin", "100mg", 0, _dt(2026, 1, 1, tz=timezone.utc), "taken"),
    ]

    result = report_service._aggregate_medicine_stats(sorted_records)

    assert len(result) == 2
    dosages = {stat["dosage"] for stat in result}
    assert dosages == {"500mg", "100mg"}


def test_aggregate_medicine_stats_sorted_alphabetically_case_insensitive():
    sorted_records = [
        ("zebra med", "1mg", 0, _dt(2026, 1, 1, tz=timezone.utc), "taken"),
        ("Aspirin", "500mg", 0, _dt(2026, 1, 1, tz=timezone.utc), "taken"),
        ("banana med", "1mg", 0, _dt(2026, 1, 1, tz=timezone.utc), "taken"),
    ]

    result = report_service._aggregate_medicine_stats(sorted_records)

    names = [stat["name"] for stat in result]
    assert names == ["Aspirin", "banana med", "zebra med"]


def test_get_medicine_stats_summary_end_to_end():
    """Public entry point should combine prep + aggregation correctly."""
    records = [
        ("Aspirin", "500mg", 5, _dt(2026, 1, 1, 8, 0, tz=timezone.utc), "taken"),
        ("Aspirin", "500mg", 5, _dt(2026, 1, 2, 8, 0, tz=timezone.utc), "skipped"),
    ]

    result = report_service.get_medicine_stats_summary(records, user_tz="Europe/Kyiv")

    assert len(result) == 1
    assert result[0]["name"] == "Aspirin"
    assert result[0]["total"] == 2


def test_get_medicine_stats_summary_empty_records_returns_empty_list():
    assert report_service.get_medicine_stats_summary([]) == []


# get_clean_status_text

def test_get_clean_status_text_strips_emojis():
    taken_text = report_service._get_clean_status_text("taken", "ua")
    skipped_text = report_service._get_clean_status_text("skipped", "ua")

    for text in (taken_text, skipped_text):
        assert "✅" not in text
        assert "❌" not in text
        assert "⏭️" not in text


# create_excel_report / create_csv_report - smoke tests

def _sample_records():
    return [
        ("Aspirin", "500mg", 5, _dt(2026, 1, 1, 8, 0, tz=timezone.utc), "taken"),
        ("Aspirin", "500mg", 5, _dt(2026, 1, 2, 8, 0, tz=timezone.utc), "skipped"),
        ("Ibuprofen", "200mg", 3, _dt(2026, 1, 3, 20, 0, tz=timezone.utc), "taken"),
    ]


def test_create_excel_report_returns_non_empty_buffer():
    buffer = report_service.create_excel_report(_sample_records(), lang="ua", user_name="Redi")

    assert buffer.getbuffer().nbytes > 0


def test_create_excel_report_is_valid_workbook_with_three_sheets():
    buffer = report_service.create_excel_report(_sample_records(), lang="ua", user_name="Redi")

    wb = openpyxl.load_workbook(buffer)
    assert len(wb.sheetnames) == 3


def test_create_excel_report_data_rows_match_record_count():
    records = _sample_records()
    buffer = report_service.create_excel_report(records, lang="ua", user_name="Redi")

    wb = openpyxl.load_workbook(buffer)
    ws = wb[wb.sheetnames[0]]
    data_rows = [row for row in ws.iter_rows(min_row=4, values_only=True) if row[0] is not None]
    assert len(data_rows) == len(records)


def test_create_excel_report_handles_empty_records_without_crashing():
    buffer = report_service.create_excel_report([], lang="ua", user_name="Redi")

    assert buffer.getbuffer().nbytes > 0
    wb = openpyxl.load_workbook(buffer)
    assert len(wb.sheetnames) == 3


def test_create_csv_report_returns_non_empty_buffer():
    buffer = report_service.create_csv_report(_sample_records(), lang="ua", user_name="Redi")

    assert buffer.getbuffer().nbytes > 0


def test_create_csv_report_contains_one_row_per_record():
    records = _sample_records()
    buffer = report_service.create_csv_report(records, lang="ua", user_name="Redi")

    content = buffer.getvalue().decode("utf-8")
    lines = [line for line in content.splitlines() if line.strip()]
    assert len(lines) == 2 + len(records)


def test_create_csv_report_handles_empty_records_without_crashing():
    buffer = report_service.create_csv_report([], lang="ua", user_name="Redi")

    assert buffer.getbuffer().nbytes > 0
